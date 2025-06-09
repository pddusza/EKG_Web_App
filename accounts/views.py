import os, io, uuid
from django.conf                 import settings
from django.shortcuts            import render, redirect, get_object_or_404
from django.contrib.auth         import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.forms   import AuthenticationForm, UserCreationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from .forms                      import CustomUserCreationForm, CSVResultForm, ProfileForm
from .models import CSVResult
from ecg.models            import ECGSignal, AnalysisResult
from ecg.ml                import run_ecg_analysis, predict_from_csv
import csv
import json
from django.core.paginator import Paginator
from .models                    import Profile
from PIL                 import Image
from django.core.files.base    import ContentFile
from django.contrib import messages
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.pdfbase     import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference, Series
from django.utils      import timezone
from reportlab.pdfbase      import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes   import A4
from reportlab.lib            import colors
from reportlab.platypus       import Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums      import TA_LEFT, TA_CENTER
from reportlab.pdfgen         import canvas
from .models                  import CSVResult

import math


# Register a Unicode font for Polish characters
FONT_NAME = "DejaVuSans"
FONT_PATH = os.path.join(settings.BASE_DIR, "reports", "fonts", "DejaVuSans.ttf")
pdfmetrics.registerFont(TTFont(FONT_NAME, FONT_PATH))

# Define paragraph styles
styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name='MetricNote', fontName=FONT_NAME, fontSize=8, leading=10, leftIndent=10, rightIndent=10))
styles.add(ParagraphStyle(name='Summary', fontName=FONT_NAME, fontSize=10, leading=12))

app_name = 'accounts'

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
            profile = user.profile
            if profile.is_doctor or profile.is_patient:
                return redirect('accounts:mainscreen')
            return redirect('accounts:login')
    else:
        form = AuthenticationForm()
    return render(request, 'accounts/login.html', {'form': form})


def register(request):
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()            # <-- writes to auth_user
            auth_login(request, user)
            return redirect("accounts:login")
    else:
        form = CustomUserCreationForm()

    return render(request, "accounts/register.html", {"form": form})

@login_required
def mainscreen_view(request):
    return render(request, 'accounts/mainscreen.html')

def user_logout(request):
    auth_logout(request)
    return redirect('accounts:login')


@login_required
def profile_view(request):
    user    = request.user
    profile = user.profile
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            # Update User
            user.username = form.cleaned_data['username']
            user.email    = form.cleaned_data['email']
            user.save()
            # Save Profile fields
            profile = form.save(commit=False)
            avatar_file = form.cleaned_data.get('avatar_upload')
            if avatar_file:
                profile.avatar = avatar_file
            profile.save()
            # Crop
            offsetX = form.cleaned_data['offsetX']
            offsetY = form.cleaned_data['offsetY']
            scale   = form.cleaned_data['scale']
            _crop_and_save(profile, offsetX, offsetY, scale)
            profile.save()
            return redirect('accounts:mainscreen')
    else:
        form = ProfileForm(
            initial={'username': request.user.username,
                     'email':    request.user.email,
                     'offsetX':  0,'offsetY':0,'scale':1.0},
            instance=profile
        )
    return render(request, 'accounts/profile.html', {'form': form, 'profile': profile})

def _crop_and_save(profile, offsetX, offsetY, scale):
    """
    Open profile.avatar, resize and crop to 120×120, then
    overwrite profile.avatar with a new UUID file in avatars_users.
    """
    img_path = profile.avatar.path
    with Image.open(img_path) as img:
        # 1) Resize
        new_w = int(img.width * scale)
        new_h = int(img.height * scale)
        img = img.resize((new_w, new_h), Image.LANCZOS)

        # 2) Centered crop box + offsets
        frame = 120
        cx = new_w//2 + offsetX
        cy = new_h//2 + offsetY
        left   = max(0, cx - frame//2)
        top    = max(0, cy - frame//2)
        right  = min(new_w, left + frame)
        bottom = min(new_h, top  + frame)
        cropped = img.crop((left, top, right, bottom))

        # 3) Write into buffer
        buf = io.BytesIO()
        cropped.save(buf, format='PNG')
        buf.seek(0)

        # 4) New UUID file path
        fname = f"{uuid.uuid4().hex}.png"

        # 5) Remove old file & save new one
        profile.avatar.delete(save=False)
        profile.avatar.save(fname,ContentFile(buf.read()),save=False)




def settings_view(request):
    return render(request, 'accounts/settings.html')

def your_results_view(request):

    return render(request, 'accounts/your_results.html')

def account_type_view(request):
    return render(request, 'accounts/account_type.html')

@login_required
def choose_patient_view(request):
    if request.method == 'POST':
        pesel = request.POST.get('pesel')

        try:
            profile = Profile.objects.get(pesel=pesel, is_patient=True)
        except Profile.DoesNotExist:
            messages.error(request, "Pacjent o podanym PESELu nie istnieje.")
            return redirect('accounts:choose_patient')

        if profile.doctor is not None:
            messages.warning(request, "Pacjent jest już przypisany do innego lekarza.")
            return redirect('accounts:choose_patient')

        profile.doctor = request.user
        profile.save()
        messages.success(request, f"Pacjent z PESEL {pesel} został przypisany do Ciebie.")
        return redirect('accounts:my_patients')  

    return render(request, 'accounts/choose_patient.html')

def register_patient_view(request):
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            # 1) Create user & get profile
            user    = form.save()
            profile = user.profile

            # 2) Populate patient fields...
            profile.first_name      = form.cleaned_data['first_name']
            profile.last_name       = form.cleaned_data['last_name']
            profile.pesel           = form.cleaned_data['pesel']
            profile.birth_date      = form.cleaned_data['birth_date']
            profile.medical_history = form.cleaned_data['medical_history']
            profile.save()

            # 3) Determine raw avatar content: upload or stock
            upload = form.cleaned_data['avatar_upload']
            choice = form.cleaned_data['avatar_choice']
            if upload:
                raw = upload.read()
                ext = os.path.splitext(upload.name)[1]
            else:
                # stock path must include the avatars folder
                stock_path = os.path.join(settings.MEDIA_ROOT, 'avatars', choice)
                with open(stock_path, 'rb') as f:
                    raw = f.read()
                ext = os.path.splitext(choice)[1]

            # 2) Save raw into avatars_users/<uuid>.ext
            fname = f"{uuid.uuid4().hex}{ext}"
            profile.avatar.save(fname, ContentFile(raw), save=False)
            profile.save()

            # 3) Crop per hidden inputs
            try:
                scale   = float(request.POST.get('scale', 1.0))
                offsetX = int(request.POST.get('offsetX', 0))
                offsetY = int(request.POST.get('offsetY', 0))
            except ValueError:
                scale, offsetX, offsetY = 1.0, 0, 0

            _crop_and_save(profile, offsetX, offsetY, scale)
            profile.save()

            auth_login(request, user)
            return redirect("accounts:login")
    else:
        form = CustomUserCreationForm()
    return render(request, "accounts/register_patient.html", {"form": form})


def register_doctor_view(request):
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            # 1) Create user & get profile
            user    = form.save()
            profile = user.profile

            # 2) Populate doctor fields...
            profile.first_name     = form.cleaned_data['first_name']
            profile.last_name      = form.cleaned_data['last_name']
            profile.license_number = form.cleaned_data['license_number']
            profile.bio            = form.cleaned_data['bio']
            # mark roles
            profile.is_patient = False
            profile.is_doctor  = True
            profile.is_admin   = True
            profile.save()

            # 3) Determine raw avatar bytes
            upload = form.cleaned_data['avatar_upload']
            choice = form.cleaned_data['avatar_choice']
            if upload:
                raw = upload.read()
                ext = os.path.splitext(upload.name)[1]
            else:
                stock_path = os.path.join(settings.MEDIA_ROOT, 'avatars', choice)
                with open(stock_path, 'rb') as f:
                    raw = f.read()
                ext = os.path.splitext(choice)[1]

            fname = f"{uuid.uuid4().hex}{ext}"
            profile.avatar.save(fname, ContentFile(raw), save=False)
            profile.save()

            try:
                scale   = float(request.POST.get('scale', 1.0))
                offsetX = int(request.POST.get('offsetX', 0))
                offsetY = int(request.POST.get('offsetY', 0))
            except ValueError:
                scale, offsetX, offsetY = 1.0, 0, 0

            _crop_and_save(profile, offsetX, offsetY, scale)
            profile.save()

            auth_login(request, user)
            return redirect("accounts:login")
    else:
        form = CustomUserCreationForm()
    return render(request, "accounts/register_doctor.html", {"form": form})


def _sanitize_for_json(obj):
    """
    Recursively replace any float that is NaN/Inf with None,
    so that the resulting structure is valid JSON.
    """
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    if isinstance(obj, float):
        return obj if math.isfinite(obj) else None
    return obj

@login_required
@user_passes_test(lambda u: u.profile.is_doctor)
def add_result_view(request):
    if request.method == 'POST':
        form = CSVResultForm(request.POST, request.FILES)
        if form.is_valid():
            # --- 1) find patient by PESEL under this doctor ---
            pesel = form.cleaned_data['pesel']
            try:
                profile = Profile.objects.get(
                    pesel=pesel,
                    is_patient=True,
                    doctor=request.user
                )
                patient_user = profile.user
            except Profile.DoesNotExist:
                form.add_error('pesel', "Pacjent o podanym numerze PESEL nie istnieje.")
                return render(request, 'accounts/add_result.html', {'form': form})

            # --- 2) save the CSVResult for that patient ---
            result = form.save(commit=False)
            result.owner = patient_user
            result.save()

            # --- 3) run ECG analysis & classification ---
            raw_stats = run_ecg_analysis(result.csv_file.path)
            try:
                classification = predict_from_csv(result.csv_file.path)
            except Exception as e:
                classification = {'error': str(e)}
            raw_stats['classification'] = classification

            # sanitize NaN/Inf → None
            def _sanitize(obj):
                if isinstance(obj, dict):
                    return {k:_sanitize(v) for k,v in obj.items()}
                if isinstance(obj, list):
                    return [_sanitize(v) for v in obj]
                if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
                    return None
                return obj

            clean_stats = _sanitize(raw_stats)
            result.analysis = clean_stats
            result.save()

            # mirror into ecg app
            ecg_signal = ECGSignal.objects.create(
                owner=patient_user,
                file=result.csv_file
            )
            AnalysisResult.objects.create(
                signal=ecg_signal,
                result_json=clean_stats
            )

            return redirect('accounts:your_results')
    else:
        form = CSVResultForm()

    return render(request, 'accounts/add_result.html', {'form': form})

@login_required
def your_results_view(request):
    # default owner is the logged-in user
    owner = request.user

    # if doctor is viewing a specific patient
    patient_id = request.GET.get('patient_id')
    if patient_id and request.user.profile.is_doctor:
        try:
            profile = Profile.objects.get(
                user__id=patient_id,
                is_patient=True,
                doctor=request.user
            )
            owner = profile.user
        except Profile.DoesNotExist:
            messages.error(request, "Nie masz dostępu do wyników tego pacjenta.")
            return redirect('accounts:my_patients')

    qs = CSVResult.objects.filter(owner=owner).order_by('-uploaded_at')

    name_filter = request.GET.get('name')
    if name_filter:
        qs = qs.filter(title__icontains=name_filter)

    # inject ML stats exactly as before
    results = []
    for r in qs:
        try:
            ecg  = ECGSignal.objects.get(file=r.csv_file.name, owner=owner)
            ar   = AnalysisResult.objects.filter(signal=ecg).order_by('-id').first()
            stats = ar.result_json if ar else None
        except ECGSignal.DoesNotExist:
            stats = None
        r.ml_stats = stats
        results.append(r)

    # record_number override
    record_number = request.GET.get('record_number')
    if record_number:
        try:
            index = int(record_number) - 1
            if 0 <= index < len(results):
                results = [results[index]]
            else:
                results = []
        except ValueError:
            results = []
        return render(request, 'accounts/your_results.html', {
            'results': results,
            'page_obj': None,
            'page': None
        })

    # pagination (3 per page)
    paginator = Paginator(results, 3)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    offset    = (page_obj.number - 1) * 3

    return render(request, 'accounts/your_results.html', {
        'results':      page_obj.object_list,
        'page_obj':     page_obj,
        'current_page': page_obj.number,
        'offset':       offset
    })


@login_required
def result_detail_view(request, pk):
    """
    Patients may only see their own results.
    Doctors may see results for any patient assigned to them.
    """
    user = request.user
    # Determine which CSVResult to fetch
    if hasattr(user, 'profile') and user.profile.is_doctor:
        # Doctor: allow if result.owner.profile.doctor == request.user
        result = get_object_or_404(
            CSVResult,
            pk=pk,
            owner__profile__doctor=user
        )
    else:
        result = get_object_or_404(CSVResult, pk=pk, owner=user)

    # Load raw samples for the signal chart
    samples = []
    try:
        with open(result.csv_file.path, newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                try:
                    samples.append(float(row[0]))
                except ValueError:
                    pass
    except Exception:
        samples = []

    signal_json = json.dumps(samples)

    try:
        ecg_signal = ECGSignal.objects.get(file=result.csv_file.name, owner=result.owner)
        ar = AnalysisResult.objects.filter(signal=ecg_signal).order_by('-id').first()
        analysis = ar.result_json if ar else {}
    except ECGSignal.DoesNotExist:
        analysis = {}

    # Unpack fields for template
    stats          = analysis.get('stats', {})
    classification = analysis.get('classification', {})
    rpeaks         = analysis.get('rpeaks', [])
    heart_rate     = analysis.get('heart_rate', [])
    hrv_time       = analysis.get('hrv_time', {})
    hrv_frequency  = analysis.get('hrv_frequency', {})
    hrv_nonlinear  = analysis.get('hrv_nonlinear', {})
    morphology     = analysis.get('morphology', {})

    return render(request, 'accounts/result_detail.html', {
        'result':       result,
        'signal_json':  signal_json,
        'stats':        stats,
        'rpeaks':       rpeaks,
        'heart_rate':   heart_rate,
        'hrv_time':     hrv_time,
        'hrv_frequency':hrv_frequency,
        'hrv_nonlinear':hrv_nonlinear,
        'morphology':   morphology,
        'classification': classification,
    })

def is_doctor(user):
    return hasattr(user, 'profile') and user.profile.is_doctor

def is_patient(user):
    return hasattr(user, 'profile') and user.profile.is_patient


@login_required
def my_patients_view(request):
    pesel_filter = request.GET.get('pesel', '')
    # only patients assigned to this doctor
    qs = Profile.objects.filter(
        is_patient=True,
        doctor=request.user
    )
    if pesel_filter:
        qs = qs.filter(pesel__icontains=pesel_filter)
    return render(request, 'accounts/my_patients.html', {
        'patients':    qs,
        'pesel_filter': pesel_filter
    })

\
@login_required
def download_my_results_xlsx(request, pk):
    # — Permission logic —
    if hasattr(request.user, 'profile') and request.user.profile.is_doctor:
        result = get_object_or_404(
            CSVResult, pk=pk,
            owner__profile__doctor=request.user
        )
    else:
        result = get_object_or_404(CSVResult, pk=pk, owner=request.user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Wyniki i surowe dane"
    row = 1

    # 1) Metadata + Średnie HR
    hr_list = (result.analysis or {}).get("heart_rate", [])
    avg_hr   = sum(hr_list)/len(hr_list) if hr_list else None
    meta = [
        ("Tytuł",      result.title),
        ("PESEL",      result.pesel),
        ("Data",       result.exam_date.isoformat()),
        ("Średnie HR", avg_hr),
    ]
    for label, val in meta:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=val)
        if label == "Średnie HR" and avg_hr is not None:
            if avg_hr < 90:
                color = "FFCCEECC"
            elif avg_hr <= 120:
                color = "FFFFF2CC"
            else:
                color = "FFF4CCCC"
            cell.fill = PatternFill("solid", fgColor=color)
        row += 1
    row += 1

    # 2) Raw CSV marker
    start_raw = row
    ws.cell(row=row, column=1, value="--- Surowe dane CSV ---")
    row += 1

    # 3) Dump CSV with numeric casting
    with result.csv_file.open("rb") as f:
        text_stream = io.TextIOWrapper(f, encoding="utf-8", newline="")
        reader = csv.reader(text_stream, delimiter=',')
        # get first row to count leads
        try:
            first_row = next(reader)
        except StopIteration:
            first_row = []
        num_leads = len(first_row)

        # write header
        ws.cell(row=row, column=1, value="Index")
        for i in range(1, num_leads+1):
            ws.cell(row=row, column=1+i, value=f"Lead_{i}")
        row += 1

        # write first data row
        idx = 1
        ws.cell(row=row, column=1, value=idx)
        for col_idx, raw in enumerate(first_row, start=2):
            try:
                val = float(raw)
            except (ValueError, TypeError):
                val = None
            ws.cell(row=row, column=col_idx, value=val)
        row += 1
        idx += 1

        # write rest
        for data_row in reader:
            ws.cell(row=row, column=1, value=idx)
            for col_idx, raw in enumerate(data_row, start=2):
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    val = None
                ws.cell(row=row, column=col_idx, value=val)
            row += 1
            idx += 1

    # 4) Build chart using header row for series names
    header_row     = start_raw + 1
    first_data_row = header_row + 1
    last_data_row  = row - 1
    first_lead_col = 2
    last_lead_col  = first_lead_col + num_leads - 1

    chart = LineChart()
    chart.title        = "Sygnały ECG"
    chart.x_axis.title = "Index"
    chart.y_axis.title = "Amplitude"

    data = Reference(
        ws,
        min_col=first_lead_col,
        max_col=last_lead_col-10,
        min_row=header_row,
        max_row=last_data_row
    )
    chart.add_data(data, titles_from_data=True)

    cats = Reference(
        ws,
        min_col=1,
        min_row=first_data_row,
        max_row=last_data_row
    )
    chart.set_categories(cats)

    insert_col = last_lead_col + 2
    ws.add_chart(chart, f"{chr(64 + insert_col)}{start_raw}")

    # 5) Stream back
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    resp = HttpResponse(
        out.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )
    resp["Content-Disposition"] = 'attachment; filename="wyniki_full.xlsx"'
    return resp










@login_required
def download_pdf(request, pk):
    # Permission logic
    if hasattr(request.user, 'profile') and request.user.profile.is_doctor:
        result = get_object_or_404(
            CSVResult,
            pk=pk,
            owner__profile__doctor=request.user
        )
    else:
        result = get_object_or_404(CSVResult, pk=pk, owner=request.user)

    data       = result.analysis or {}
    exam_date  = result.uploaded_at.date().strftime("%Y-%m-%d")
    issue_date = timezone.localdate().strftime("%Y-%m-%d")

    buffer = io.BytesIO()
    p      = canvas.Canvas(buffer, pagesize=A4)
    w, h    = A4

    # Header
    p.setFont(FONT_NAME, 16)
    p.drawString(50, h - 50, f"Wynik: {result.title}")
    p.setFont(FONT_NAME, 10)
    p.drawString(50, h - 70, f"Data badania: {exam_date}")
    p.drawRightString(w - 50, h - 70, f"Data wydania: {issue_date}")

    # Classification table
    cls    = data.get('classification', {})
    label  = (cls.get('predictions') or ["–"])[0]
    interp = "Normalny sygnał" if label == "NORM" else "Skonsultuj się ze swoim lekarzem"
    tbl1 = Table(
        [["Klasyfikacja", "Interpretacja"], [label, interp]],
        colWidths=[200, 300]
    )
    tbl1.setStyle(TableStyle([
        ('FONT',       (0,0), (-1,-1), FONT_NAME),
        ('FONTSIZE',   (0,0), (-1,-1), 12),
        ('ALIGN',      (0,0), (-1,0), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
    ]))
    w1, h1 = tbl1.wrapOn(p, w, h)
    tbl1.drawOn(p, 50, h - 100 - h1)
    cursor_y = h - 100 - h1 - 20

    # Metrics list with notes
    hr_list = data.get("heart_rate", [])
    hr_mean = sum(hr_list)/len(hr_list) if hr_list else None
    metrics = [
        ("Średnie HR (bpm)", hr_mean,      60,   100,   "between",
         "Norma 60–100 bpm; wskaźnik sprawności układu krążenia."),
        ("PR (ms)",          data.get("morphology", {}).get("pr_mean_ms"), 120, 200, "between",
         "Prawidłowy PR: 120–200 ms; dłuższy może wskazywać na blok AV."),
        ("QRS (ms)",         data.get("morphology", {}).get("qrs_mean_ms"), 70,  100, "between",
         "Norma QRS: 70–100 ms; szerszy niż 120 ms oznacza szeroki QRS."),
        ("QT (ms)",          data.get("morphology", {}).get("qt_mean_ms"),  350, None, "lt",
         "Norma QT: <350 ms (QTc <440 ms mężczyźni, <460 ms kobiety)."),
        ("SDNN (ms)",        data.get("hrv_time", {}).get("HRV_SDNN"),     50,   None, "gt",
         "SDNN ≥50 ms – prawidłowa zmienność; <50 ms – podwyższone ryzyko."),
        ("RMSSD (ms)",       data.get("hrv_time", {}).get("HRV_RMSSD"),    19,   None, "gt",
         "RMSSD 19–107 ms typowe u dorosłych; <19 ms może wskazywać na autonomiczną dysfunkcję."),
        ("pNN50 (%)",        data.get("hrv_time", {}).get("HRV_pNN50"),    3,    None, "gt",
         "pNN50 >3 % – norma; <3 % – obniżona zmienność, ryzyko arytmii."),
    ]

    # Build metrics table data
    tbl2_data = [["Parametr", "Wartość", "Norma", "Status"]]
    abnormal = []
    def check(val, low, high, cmp):
        if val is None:
            return "–", "nieobliczone"
        if cmp == "between": ok = low <= val <= high
        elif cmp == "lt":     ok = val < low
        elif cmp == "gt":     ok = val > low
        else:                   ok = True
        return f"{val:.1f}", ("w normie" if ok else "poza normą")

    for name, val, low, high, cmp, note in metrics:
        disp, status = check(val, low, high, cmp)
        rng = f"{low}–{high}" if high else ("<"+str(low) if cmp=="lt" else ">"+str(low))
        tbl2_data.append([name, disp, rng, status])
        if status == "poza normą":
            abnormal.append((name, note))

    # Draw metrics table
    tbl2 = Table(tbl2_data, colWidths=[150, 80, 100, 180])
    tbl2.setStyle(TableStyle([
        ('FONT',     (0,0),(-1,-1), FONT_NAME),
        ('FONTSIZE', (0,0),(-1,-1), 10),
        ('GRID',     (0,0),(-1,-1), 0.3, colors.black),
        ('BACKGROUND',(0,0),(-1,0), colors.lightgrey),
        ('VALIGN',   (0,0),(-1,-1), 'TOP'),
    ]))
    w2, h2 = tbl2.wrapOn(p, w, h)
    tbl2.drawOn(p, 50, cursor_y - h2)
    cursor_y -= (h2 + 10)

    # Draw metric notes under table, each as a paragraph
    for nm, note in abnormal:
        para = Paragraph(f"<b>{nm}</b>: {note}", styles['MetricNote'])
        w3, h3 = para.wrap(w-100, h)
        para.drawOn(p, 50, cursor_y - h3)
        cursor_y -= (h3 + 5)

    # --- Dynamic Summary ---
    summary_parts = []
    if not abnormal:
        summary_text = "Wszystkie metryki mieszczą się w normie."
    else:
        metrics_list = ", ".join([nm for nm, _ in abnormal[:3]])
        tail = f" i {len(abnormal)-3} inne." if len(abnormal)>3 else "."
        summary_text = f"Parametry poza normą: {metrics_list}{tail} Proszę skonsultować się z lekarzem."

    para_sum = Paragraph(summary_text, styles['Summary'])
    w4, h4 = para_sum.wrap(w-100, h)
    para_sum.drawOn(p, 50, cursor_y - h4)

    # Finish up
    p.showPage()
    p.save()
    buffer.seek(0)

    filename = f"Wyniki_{request.user.first_name}_{request.user.last_name}_{exam_date}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



